from datetime import date, datetime
import sys
from openpyxl.utils import get_column_letter
import os #make_responce pour envoye des fichier a telecharger 
import pandas as pd #pour creation des tableau excel
from io import BytesIO # type: ignore #pour la cration du fichier excel
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from werkzeug.security import generate_password_hash
from flask import Flask,current_app, render_template_string, request, render_template, redirect,Response,jsonify, url_for, flash, session,make_response # type: ignore
from flask_sqlalchemy import SQLAlchemy # type: ignore
from werkzeug.security import check_password_hash # type: ignore
from werkzeug.utils import secure_filename
import random
import time
from flask_mail import Mail, Message # type: ignore
from recognition.Reco import encoding, systeme_Recognition,test_camera
from recognition.encoding import update_json_data
import cv2 # type: ignore
from datetime import datetime
app=Flask(__name__)
# Configuration Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'presenciaApp2025@gmail.com'         # ← Ton email
app.config['MAIL_PASSWORD'] = 'xfey cjkq wxvr nlke'   # ← Mot de passe ou mot de passe d'application 
mail = Mail(app)
app.secret_key = "cle_secrete"
# Connexion à MySQL
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+mysqlconnector://root:@localhost/pfe'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
camera=None
active=False
dernier_noms_detectes = []
known_encodings, known_names = encoding()
dernier_message_reconnaissance = ""
derniere_detection_stable = None
temps_detection_stable = 0
duree_detection_requise = 3
# Modèle Admin
class Admin(db.Model):
    __tablename__ = 'admins'  # Nom de la table déjà existante
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    email = db.Column(db.String(400), nullable=False)  # ← important
    photo = db.Column(db.String(255),default='profil.jpg')  # chemin ou nom de fichier image
    created_at = db.Column(db.DateTime,default=datetime.now().date)
    last_login = db.Column(db.DateTime, nullable=True)
class Employes(db.Model):
    __tablename__='employes'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    matricule = db.Column(db.String(20), unique=True, nullable=False)
    nom = db.Column(db.String(50), nullable=False)
    prenom = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    nom_complet = db.Column(db.String(200), nullable=True)
    telephone = db.Column(db.String(20), nullable=True)
    poste = db.Column(db.String(50), nullable=True)
    date_embauche = db.Column(db.Date, nullable=True)
    statut = db.Column(db.String(20), nullable=False, default='actif')
    photo = db.Column(db.String(255),default='profil.jpg')  # chemin ou nom de fichier image
class Presence(db.Model):
    __tablename__ = 'presence'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    employe_id = db.Column(db.Integer, db.ForeignKey('employes.id'), nullable=False)
    date_presence = db.Column(db.Date, nullable=False)
    heure_entree = db.Column(db.Time, nullable=False)
    statut = db.Column(db.String(20))  # Exemples : 'présent', 'absent', 'en retard'
     # Exemples : 'manuel', 'reconnaissance faciale'
    mode = db.Column(db.String(255)) 
    commentaire = db.Column(db.String(255),default='-----')  
    # Relation avec l'employé (si tu veux accéder à employe depuis presence)
    employe = db.relationship('Employes', backref='presences')
class Activity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(255), nullable=False)
    date_action = db.Column(db.DateTime, default=datetime.now)
def save_activity(description):
    new_activity = Activity(
        description=description,
        date_action=datetime.now()  # ici tu peux aussi utiliser un helper pour fuseau horaire si besoin
    )
    db.session.add(new_activity)
    db.session.commit()
def est_deja_enregistre(employe_id):
    aujourd_hui = date.today()
    presence = Presence.query.filter_by(
        employe_id=employe_id,
        date_presence=aujourd_hui
    ).first()
    return presence is not None
def get_infos(nom):
    employe=Employes.query.filter_by(nom_complet=nom).first()
    if employe:
        return  {
            "matricule": employe.matricule,
            "prenom": employe.prenom,
            "nom": employe.nom,
            "poste": employe.poste,
        }
    else:
        return None
admin=None
temps_dernier_visage = time.time()
@app.route("/")
def home():
    
    return render_template('home.html',show_login_popup=False) 
@app.route('/login', methods=['POST'])
def login():
        utilisateur = request.form['username']
        mot_de_passe = request.form['password']

        # Vérification simple
        user = Admin.query.filter_by(username=utilisateur).first()
        if user and check_password_hash(user.password, mot_de_passe):
           admin_id=session['admin_id'] = user.id
           admin = Admin.query.get(admin_id)
           session['admin'] =utilisateur
           admin.last_login = datetime.now()
           db.session.commit()
           return render_template('home.html', success=True)
        else:
           flash("Nom ou mot de passe incorrect", "danger")
        # On indique à la page qu’il faut ouvrir la pop-up
           return render_template('home.html', show_login_popup=True)
@app.route('/refresh_json',methods=['POST'])
def refresh_json():
    global known_encodings, known_names 
    update_json_data()
    known_encodings, known_names = encoding()
    save_activity("Mise à jour du fichier JSON")
    return render_template_string("""
        <script>
            alert(' Le fichier encodages.json a été mis à jour avec succès.');
            window.location.href = '/Admin';  
        </script>
    """)
@app.route('/update_admin', methods=['POST'])
def update_admin():
    admin_id = session.get('admin_id')
    admin = Admin.query.get(admin_id)
    if request.method == 'POST':
        username = request.form.get('nom_complet')
        email = request.form.get('email')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        if new_password and new_password != confirm_password:
            return render_template_string("""
                <script>
                    alert('Les mots de passe ne correspondent pas');
                    window.location.href = '/Admin';
                </script>
            """)
        admin.username = username
        admin.email = email
        if new_password:
            admin.password = generate_password_hash(new_password)

        db.session.commit()
        save_activity("Modification des informations de l'administrateur")
        return render_template_string("""
            <script>
                alert('Données modifiées avec succès');
                window.location.href = '/Admin';
            </script>
        """)
@app.route('/add_admin', methods=['POST'])
def add_admin():
    fullname = request.form.get('fullname')
    email = request.form.get('email')
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    if password != confirm_password:
        return render_template_string("""
             <script>
                    alert('Les mots de passe ne correspondent pas');
                    window.location.href = '/Admin';
                </script>
            """)
    hashed_password = generate_password_hash(password)
    new_admin = Admin(username=fullname, email=email, password=hashed_password)
    try:
        db.session.add(new_admin)
        db.session.commit()
        flash("Administrateur ajouté avec succès.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de l'ajout : {str(e)}", "error")

    return redirect(url_for('admin'))
@app.route('/deconnexion')
def deconnexion():
    save_activity("Déconnexion de l'administrateur")
    return render_template ('home.html')
@app.route('/principale')
def principale():
    admin_id=session['admin_id']
    admin = Admin.query.get(admin_id) 
    return render_template ('acceuil.html',admin=admin)
@app.route('/propos')
def test():
    return render_template ('propos.html')
@app.route('/Admin')
def admin():
    activities = Activity.query.order_by(Activity.date_action.desc()).all()
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)
    return render_template ('Admin.html',admin=admin,activities=activities)
    
@app.route("/photoAdmin", methods=["POST"])
def photoAdmin():
    admin = Admin.query.get(session['admin_id'])
    if 'photo' in request.files:
        file = request.files['photo']
        if file.filename != "":
            from werkzeug.utils import secure_filename # type: ignore
            filename = secure_filename(file.filename)
            file.save(os.path.join("static/images", filename))
            admin.photo = filename
            db.session.commit()
            save_activity("Mise à jour de la photo de profil de l'administrateur")
    return redirect(url_for('admin'))
@app.route('/acceuil')
def acceuil():
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)
    return render_template ('acceuil.html',admin=admin)

@app.route('/reconnaissance')
def reconnaissance():
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)
    return render_template('reconnaissance.html',admin=admin)
@app.route('/start')
def start():
    global camera,active,nombre_visage
    if not active:
        camera=cv2.VideoCapture(0)
        active=True 
    return redirect(url_for('reconnaissance'))
@app.route('/stop')
def stop():
    global camera, active,dernier_message_reconnaissance
    if active and camera is not None:
        camera.release()
        active = False
        dernier_message_reconnaissance=""
    return redirect(url_for('reconnaissance'))
nombre=0
@app.route('/listePresence')
def lsitePresence():
    admin_id = session.get("admin_id")
    today = date.today().isoformat()
    admin = Admin.query.get(admin_id)
    Presences=Presence.query.filter_by(date_presence=today).all()
    date_aujourdhui = datetime.now().strftime("%d %B %Y")
    return render_template('listePresence.html',presences=Presences,date_selectionnee=today,admin=admin,date_aujourdhui=date_aujourdhui)
@app.route('/get_message')
def get_message():
    global dernier_message_reconnaissance
    return jsonify({"message": dernier_message_reconnaissance})

def enregistrer_presence(nom):
    with app.app_context():
        employe = Employes.query.filter_by(nom_complet=nom).first()
        if not employe :
            print(f"Employé {nom} non trouvé.")
            return False

        if est_deja_enregistre(employe.id):
            print(f"Présence déjà enregistrée pour {nom} aujourd'hui.")
            return False

        nouvelle_presence = Presence(
            employe_id=employe.id,
            date_presence=date.today(),
            heure_entree=datetime.now().time(),
            statut='présent',
            mode='reconnaissance faciale',
        )

        try:
            db.session.add(nouvelle_presence)
            db.session.commit()
            print(f"Présence enregistrée pour {nom}.")
            return True
        except Exception as e:
            print(f"Erreur lors de l'enregistrement: {e}")
            db.session.rollback()
            return False

def reconnaissance_facial():
    global camera, active, nombre, dernier_noms_detectes, known_encodings, known_names, dernier_message_reconnaissance, derniere_detection_stable, temps_dernier_visage
    if not active:
        print("Erreur : Impossible d'ouvrir la caméra.")
        return

    while active:
        success, frame = camera.read()
        if not success or frame is None:
            print("Aucune frame capturée.")
            continue
        
        frame, nom_detecte, nb_visage = systeme_Recognition(frame, known_encodings, known_names)  # type: ignore
        dernier_noms_detectes = nom_detecte
        frame_courante = frame.copy()
        if nb_visage == 1 and nom_detecte[0]!="Inconnu":
            nom = nom_detecte[0]  # récupère le nom détecté
            if nom == derniere_detection_stable:
                # Même visage détecté que précédemment
                if time.time() - temps_dernier_visage >= 3:
                    if enregistrer_presence(nom):
                        dernier_message_reconnaissance = f"Présence enregistrée pour {nom}."
                        freeze_until = time.time() + 2
                    else:
                        dernier_message_reconnaissance = f"Présence déjà enregistrée pour {nom} aujourd’hui."
                        temps_dernier_visage = time.time()
                else:
                    dernier_message_reconnaissance = f"Merci de rester immobile, détection de {nom} en cours de confirmation..."
            else:
                # Nouveau visage détecté, reset timer
                derniere_detection_stable = nom
                temps_dernier_visage = time.time()
                
        elif nom_detecte[0]=="Inconnu":
             dernier_message_reconnaissance = ""
        elif nb_visage > 1:
            dernier_message_reconnaissance = "Plusieurs visages détectés, veuillez vous placer seul devant la caméra."
            temps_dernier_visage = time.time()
        else:
            if time.time() - temps_dernier_visage > 2 :
                dernier_message_reconnaissance = "Aucun visage détecté."
    

        ret, buffer = cv2.imencode('.jpg', frame)
        frame_bytes = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n\r\n')

        time.sleep(0.05)  # Petite pause pour éviter saturation CPU

    # Libération caméra quand la boucle se termine
    camera.release()  

def recon():
     # 0 pour la caméra par défaut

     while True:
        # Capture une image
        success, frame = cap.read()   # type: ignore
        if not success:
            break
        else:
            # Convertit l'image en format JPEG pour l'envoyer au navigateur
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()

            # Retourne l'image sous forme de flux
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')

     cap.release()  # type: ignore # Libère la caméra lorsque terminé
@app.route('/video_feed')
def video_feed():
    return Response(recon(),mimetype='multipart/x-mixed-replace; boundary=frame')   
@app.route('/video')
def video():
    return Response( reconnaissance_facial(), mimetype='multipart/x-mixed-replace; boundary=frame')
   
@app.route('/Historique')
def Historique():
    global admin
    return render_template('Historique.html',admin=admin)
@app.route('/oublier')
def oublier():
    return render_template('mot-de-passe.html')
code_saisi=""
email_actuel=''
@app.route('/verification-email', methods=['POST'])
def verifier():
    global email_actuel
    gmail = request.form['email'] #recuperation
    print(f"Email reçu : '{gmail}'")
    gmail = gmail.strip().lower()
    admins=Admin.query.all()
    exist= False
    for admin in admins:
        print("email de la base de donne",admin.email)
        if admin.email.strip().lower() == gmail:
            exist= True
            break
    if exist:
        print("Email trouvé dans la base de données")
        email_actuel=gmail
        return redirect(url_for('envoie', destinataire=gmail))  # Redirige vers la page d'accueil
    else:
        print("Email non trouvé dans la base de données") # Message d'erreur
        return redirect(url_for('erreur'))
def generer_code():
    return str(random.randint(100000,999999))
@app.route('/erreur')
def erreur():
    return render_template('erreur.html')
@app.route('/test1')
def test1():
    return render_template('test1.html')
def generer_code_verification():
    return ''.join(random.choices('0123456789', k=6))
@app.route('/envoie/<destinataire>')
def envoie(destinataire):
    global code_saisi
    msg = Message(subject='Code de vérification pour la réinitialisation de mot de passe',
                  recipients=[destinataire], sender=('PRESENCIA','presenciApp2025@gmail.com'))
    code= generer_code_verification()
    code_saisi=code
    msg.html ="""
       <p>Bonjour,<br>
       Vous avez demandé à réinitialiser votre mot de passe<br>
       Voici votre code de vérification : <strong>{code}</strong><br>
       Veuillez ne pas le partager avec qui que ce soit.<br>
       Si vous n’êtes pas à l’origine de cette demande, vous pouvez ignorer cet e-mail en toute sécurité.<br>
       cordialement,<br>
       L’équipe PRESENCIA<p>
         """.format(code=code)
    try:
        mail.send(msg)
        message = "Email envoyé avec succès !"
        html = f"""
         <script>
        alert("{message}");
        window.location.href = "/modification_pwd";
        </script>
           """
    except Exception as e:
        message = f"Échec de l’envoi : {str(e)}"
        html = f"""
         <script>
        alert("{message}");
        window.location.href = "/oublier";
        </script>
        """
    return Response(html, mimetype='text/html')
@app.route('/modification_pwd')
def modification_pwd():
    return render_template("modification_pwd.html")
@app.route('/verification-code', methods=['GET', 'POST'])
def verification_code():
    global code_saisi
    if request.method == 'POST':
        code= request.form.get('code')
        if code== code_saisi:
            flash("Code vérifié avec succès.")
            return render_template('changer_pwd.html')
        else:
            flash("Code incorrect. Veuillez réessayer.")
            return redirect(url_for('verification_code'))
    return render_template('modification_pwd.html')
@app.route('/changer_motdepasse', methods=['POST'])
def changer_motdepasse():
    new_pwd = request.form.get('new_password')
    confirm_pwd = request.form.get('confirm_password')
    if new_pwd != confirm_pwd:
        flash("Les mots de passe ne correspondent pas.")
        return render_template('changer_pwd.html')
    # Appel de la fonction de mise à jour
    try:
        admin = Admin.query.filter_by(email=email_actuel).first()
        # ✅ Hachage sécurisé du mot de passe
        admin.password = generate_password_hash(new_pwd) # type: ignore
        db.session.commit()
        save_activity("Changement de mot de passe de l'administrateur")
        return redirect(url_for('home', msg='success_password'))
    except Exception as e:
        db.session.rollback()
        message=f"Erreur lors de la modification : {str(e)}"
        return redirect(url_for('home', msg='error_password'))
@app.route('/Employes')
def employes():
    admin_id=session['admin_id'] 
    admin = Admin.query.get(admin_id)
    employes=Employes.query.all()
    return render_template('Employes.html',employes=employes,admin=admin)
@app.route('/supprimer/<matricule>', methods=['GET'])
def supprimer(matricule):
    emp = Employes.query.filter_by(matricule=matricule).first()
    if emp:
        db.session.delete(emp)
        db.session.commit()
        save_activity(f"Suppression de l'employé {emp.nom_complet}")
    return redirect('/Employes')
@app.route('/editer/<matricule>', methods=['GET', 'POST'])
def editer_emp(matricule):
    # ton code ici
    pass
@app.route('/export_excel')
def export_excel():
    employes = Employes.query.all()
    data = []
    for emp in employes:
        data.append({
            'Matricule': emp.matricule,
            'Nom': emp.nom,
            'Prénom': emp.prenom,
            'Email': emp.email,
            'Téléphone': emp.telephone,
            'Poste': emp.poste,
            'Date Embauche': emp.date_embauche.strftime('%Y-%m-%d') if emp.date_embauche else '',
            'Statut': emp.statut
        })
    dataPandas=pd.DataFrame(data)
    # cration de fichier excel
    output=BytesIO()
    with pd.ExcelWriter(output,engine='openpyxl') as writer:
        dataPandas.to_excel(writer, index=False, sheet_name='Employes')
        worksheet = writer.sheets['Employes']
        for i, column in enumerate(dataPandas.columns, start=1):
            max_length = max(dataPandas[column].astype(str).map(len).max(), len(column))
            column_letter = get_column_letter(i)
            worksheet.column_dimensions[column_letter].width = max_length + 2
    save_activity("Exportation de la liste des employés au format Excel")
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=employes.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response
@app.route('/export_presence')
def export_presence():
    presence =Presence.query.all()
    data = []
    for p in presence:
        data.append({
            'Id': p.id,
            'Nom complet': p.employe.nom_complet,
            'date presence': p.date_presence,
            'Heure d\'entrée': p.heure_entree,
            'Statut':p.statut,
            'Mode':p.mode,
            'Commentaire':p.commentaire
        })
    dataPandas=pd.DataFrame(data)
    # cration de fichier excel
    output=BytesIO()
    with pd.ExcelWriter(output,engine='openpyxl') as writer:
        dataPandas.to_excel(writer, index=False, sheet_name='Presences')
        worksheet = writer.sheets['Presences']
        for i, column in enumerate(dataPandas.columns, start=1):
            max_length = max(dataPandas[column].astype(str).map(len).max(), len(column))
            column_letter = get_column_letter(i)
            worksheet.column_dimensions[column_letter].width = max_length + 2
    save_activity("Exportation de la liste des présences au format Excel")
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment;filename=Liste_presences.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
@app.route('/modifier_emp', methods=['POST'])
def modifier_emp():
    matricule = request.form['matricule']
    print("Matricule reçu :", matricule) 
    employe = Employes.query.filter_by(matricule=matricule).first()
    if employe:
        employe.nom = request.form['nom']
        employe.prenom = request.form['prenom']
        employe.email = request.form['email']
        employe.telephone = request.form['telephone']
        employe.poste = request.form['poste']
        employe.date_embauche = datetime.strptime(request.form['date_embauche'], '%Y-%m-%d')
        employe.statut = request.form['statut']
        photo = request.files.get('photo')
        if photo and photo.filename != '':
            # Sécuriser le nom de fichier
            filename = secure_filename(photo.filename)
            uniqfilename =employe.nom+"_"+employe.prenom+"_"+filename
            upload_path = os.path.join(current_app.root_path, 'static', 'images', uniqfilename)
            photo.save(upload_path)
            employe.photo = uniqfilename
        db.session.commit()
        save_activity(f"Modification des informations de l'employé {employe.nom_complet} (matricule: {matricule})")
        return redirect(url_for('employes'))  # adapte selon ta route principale
    else:
        return "Employé non trouvé", 404
@app.route('/employe_infos')
def employe_infos():
    global dernier_noms_detectes
    if dernier_noms_detectes:
        nom = dernier_noms_detectes[0]
        print("Nom utilisé pour la recherche:", nom)
        infos = get_infos(nom)
        print("Infos trouvées:", infos)
        if infos and active:
            return jsonify(infos)
    return jsonify({"nom": "Inconnu", "prenom": "Inconnu", "matricule": "Inconnu", "poste": "Inconnu"})
@app.route('/ajouter', methods=['POST'])
def ajouter():
    try:
        # Récupération des données du formulaire
        matricule = request.form.get('matricule')
        nom = request.form.get('nom')
        prenom = request.form.get('prenom')
        email = request.form.get('email')
        telephone = request.form.get('telephone')
        poste = request.form.get('poste')
        date_embauche = request.form.get('date_embauche')
        statut = request.form.get('statut', 'actif')
        nom_complet = f"{prenom} {nom}"
         # Vérifier si l'employé existe déjà
        existant = Employes.query.filter((Employes.matricule == matricule) | (Employes.nom_complet == nom_complet)).first()
        if existant:
            return "<script>alert('Employé déjà existant !'); history.back();</script>"

        # Création du dossier pour la personne
        dossier_employe = f"{prenom}_{nom}"
        chemin_base = r"C:\Users\ELITEBOOK\Desktop\facerecognition\recognition\Personnes"
        chemin_dossier = os.path.join(chemin_base, dossier_employe)
        try:
            os.makedirs(chemin_dossier, exist_ok=True)
        except Exception as e:
            print(f"Erreur lors de la création du dossier : {e}")
        # Gestion de la photo
        photo_filename = None
        if 'image' in request.files:
            fichier = request.files['image']
            if fichier.filename != '':
                photo_filename = secure_filename(f"{nom}_{prenom}.jpg")
                chemin_complet = os.path.join('static', 'images', photo_filename)
                fichier.save(chemin_complet)
        
        # Enregistrement en base de données
        nouvel_employe = Employes(
            matricule=matricule,
            nom=nom,
            prenom=prenom,
            email=email,
            nom_complet=nom_complet,
            telephone=telephone,
            poste=poste,
            date_embauche=date_embauche,
            statut=statut,
            photo=photo_filename
        )
        db.session.add(nouvel_employe)
        db.session.commit()
        save_activity(f"Ajout d'un nouvel employé : {nom_complet} (matricule: {matricule})")
        return "<script>alert('Employé ajouté !'); window.location='/Employes';</script>"

    except Exception as e:
        print(str(e))
        return f"<script>alert('Erreur : {str(e)}'); history.back();</script>"
@app.route('/statistique')
def statistique():
    admin_id=session['admin_id'] 
    admin = Admin.query.get(admin_id)

    return render_template('statistique.html',admin=admin)

@app.route('/filtrage-presences', methods=['POST'])
def filtrage_presences():    
    date_choisie = request.form.get('date')
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)
    presences = Presence.query.filter_by(date_presence=date_choisie).all()
    date_obj = datetime.strptime(date_choisie, "%Y-%m-%d")
    date_aujourdhui = datetime.now().strftime("%d %B %Y")
    date_formatee = date_obj.strftime("%d %B %Y")
    return render_template('listePresence.html',presences=presences,date_selectionnee=date_choisie,admin=admin,date_aujourdhui=date_formatee)
@app.route('/modifier_commentaire', methods=['POST'])
def modifier_commentaire():
    presence_id = request.form.get('presence_id')
    nouveau_commentaire = request.form.get('commentaire')
    presence = Presence.query.get(presence_id)
    date_choisie = presence.date_presence
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)
    if presence:
        presence.commentaire = nouveau_commentaire
        db.session.commit()
        presences = Presence.query.filter_by(date_presence=date_choisie).all()
    date_obj =date_choisie
    date_formatee = datetime.now().strftime("%d %B %Y")
    return render_template('listePresence.html',presences=presences,date_selectionnee=date_choisie,admin=admin,date_aujourdhui=date_formatee)
@app.route('/supprimer_presence/<int:presence_id>')
def supprimer_presence(presence_id):
    presence = Presence.query.get_or_404(presence_id)
    date_choisie = presence.date_presence
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)
    db.session.delete(presence)
    db.session.commit()
    presences = Presence.query.filter_by(date_presence=date_choisie).all()
    date_obj =date_choisie
    date_formatee = datetime.now().strftime("%d %B %Y")
    flash("Présence supprimée avec succès", "success")
    return render_template('listePresence.html',presences=presences,date_selectionnee=date_choisie,admin=admin,date_aujourdhui=date_formatee)
@app.route('/ajouter_presence_manuellement', methods=['POST'])
def ajouter_presence_manuellement():
    employe_id = request.form['employe_id']
    date_presence = request.form['date_presence']
    heure_entree = request.form['heure_entree']
    statut = request.form['statut']
    mode = request.form['mode']
    commentaire = request.form.get('commentaire', '-----')

    nouvelle_presence = Presence(
        employe_id=employe_id,
        date_presence=date_presence,
        heure_entree=heure_entree,
        statut=statut,
        mode=mode,
        commentaire=commentaire
    )
    db.session.add(nouvelle_presence)
    db.session.commit()

    flash("Présence ajoutée avec succès", "success")
    return redirect(url_for('listePresence'))  # Assure-toi que la fonction s'appelle bien listePresence

if __name__ == '__main__':
    app.run(debug=True)